"""
Serviço para exportar dados extraídos para Excel
Gera arquivo .xlsx com dados estruturados
"""

import os
import tempfile
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Serviço responsável por exportar dados para Excel"""
    
    def __init__(self):
        """Inicializa o exportador Excel"""
        self.column_headers = [
            "Banco",
            "Agência",
            "Conta",
            "Tipo de Conta",
            "CPF/CNPJ",
            "Valor (R$)"
        ]
    
    def format_value(self, value: any) -> str:
        """
        Formata valor para exibição no Excel
        
        Args:
            value: Valor a ser formatado
            
        Returns:
            String formatada
        """
        if value is None:
            return "-"
        
        if isinstance(value, float):
            return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        return str(value)
    
    def create_workbook(self, data: List[Dict[str, any]]) -> Workbook:
        """
        Cria workbook Excel com dados
        
        Args:
            data: Lista de dicionários com dados extraídos
            
        Returns:
            Workbook do openpyxl
        """
        # Cria novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Convênios Bancários"
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Adiciona cabeçalhos
        for col_num, header in enumerate(self.column_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Mapeia campos do dicionário para colunas
        field_mapping = {
            "Banco": "banco",
            "Agência": "agencia",
            "Conta": "conta",
            "Tipo de Conta": "tipo_conta",
            "CPF/CNPJ": "cpf_cnpj",
            "Valor (R$)": "valor"
        }
        
        # Adiciona dados
        for row_num, record in enumerate(data, start=2):
            for col_num, header in enumerate(self.column_headers, 1):
                field_name = field_mapping[header]
                value = record.get(field_name)
                formatted_value = self.format_value(value)
                
                cell = ws.cell(row=row_num, column=col_num, value=formatted_value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajusta largura das colunas
        column_widths = {
            "A": 25,  # Banco
            "B": 12,  # Agência
            "C": 15,  # Conta
            "D": 15,  # Tipo de Conta
            "E": 18,  # CPF/CNPJ
            "F": 15   # Valor
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Adiciona linha de total se houver valores
        if any(record.get("valor") for record in data):
            total_row = len(data) + 3
            ws.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)
            
            total_value = sum(record.get("valor", 0) or 0 for record in data)
            ws.cell(row=total_row, column=6, value=self.format_value(total_value))
            ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        return wb
    
    def export_to_file(self, data: List[Dict[str, any]], output_path: str = None) -> str:
        """
        Exporta dados para arquivo Excel
        
        Args:
            data: Lista de dicionários com dados extraídos
            output_path: Caminho do arquivo de saída (opcional)
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        if not data:
            raise ValueError("Nenhum dado para exportar")
        
        # Cria workbook
        wb = self.create_workbook(data)
        
        # Define caminho do arquivo
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                tempfile.gettempdir(),
                f"convenio_bancario_{timestamp}.xlsx"
            )
        
        # Salva arquivo
        wb.save(output_path)
        
        return output_path
    
    def export_to_bytes(self, data: List[Dict[str, any]]) -> bytes:
        """
        Exporta dados para bytes (útil para retornar via API)
        
        Args:
            data: Lista de dicionários com dados extraídos
            
        Returns:
            Bytes do arquivo Excel
        """
        if not data:
            raise ValueError("Nenhum dado para exportar")
        
        # Cria workbook
        wb = self.create_workbook(data)
        
        # Salva em bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
